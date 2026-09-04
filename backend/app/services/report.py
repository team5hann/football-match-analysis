import csv
import io
from datetime import datetime



def _date(value: datetime | None) -> str:
    return value.isoformat() if value else "No data"


def _cell(value) -> str:
    return "No data" if value is None else str(value)


def build_report_data(match, summary, metrics, events, shots, tactical, network, cluster_roles=None) -> dict:
    cluster_roles = cluster_roles or {}
    home_name = match.home_team.name if match.home_team else "Home"
    away_name = match.away_team.name if match.away_team else "Away"
    player_rows = [
        {
            "jersey_number": metric.jersey_number,
            "team": cluster_roles.get(metric.team_color_cluster, "Unknown").title(),
            "track_id": metric.track_id,
            "touches": metric.touches,
            "distance_meters": metric.distance_meters,
            "average_speed_mps": metric.average_speed_mps,
            "max_speed_mps": metric.max_speed_mps,
        }
        for metric in metrics
    ]
    shot_rows = [
        {"time": shot.timestamp_seconds, "team": shot["team_role"], "track_id": shot["track_id"], "xg": shot["xg"]}
        for shot in shots
    ]
    tactical = tactical or {
        "home": type("Tactical", (), {"formation": "No data", "width": 0, "depth": 0, "compactness": 0})(),
        "away": type("Tactical", (), {"formation": "No data", "width": 0, "depth": 0, "compactness": 0})(),
    }
    network = network or {"home": {"edges": []}, "away": {"edges": []}}
    return {
        "overview": {
            "Home team": home_name,
            "Away team": away_name,
            "Competition": match.competition or "No data",
            "Date": _date(match.match_date),
            "Score": f"{match.home_score if match.home_score is not None else '-'} : {match.away_score if match.away_score is not None else '-'}",
            "Home possession %": summary.home_possession_pct if summary else 0,
            "Away possession %": summary.away_possession_pct if summary else 0,
            "Home xG": sum(row["xg"] for row in shot_rows if row["team"].lower() == "home"),
            "Away xG": sum(row["xg"] for row in shot_rows if row["team"].lower() == "away"),
            "Home formation": tactical["home"].formation,
            "Away formation": tactical["away"].formation,
        },
        "players": player_rows,
        "events": [
            {"time": event.timestamp_seconds, "type": event.event_type, "track_id": event.track_id, "description": event.description}
            for event in events
        ],
        "shots": shot_rows,
        "tactical": [
            {
                "team": role.title(),
                "formation": tactical[role].formation,
                "width": tactical[role].width,
                "depth": tactical[role].depth,
                "compactness": tactical[role].compactness,
            }
            for role in ("home", "away")
        ],
        "passing": [
            {"team": role.title(), "source": edge["source_track_id"], "target": edge["target_track_id"], "passes": edge["pass_count"]}
            for role in ("home", "away")
            for edge in network[role]["edges"]
        ],
    }


def render_csv(data: dict) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Section", "Field", "Value"])
    for field, value in data["overview"].items():
        writer.writerow(["Overview", field, value])
    for row in data["players"]:
        writer.writerow(["Players", "", *row.values()])
    for row in data["events"]:
        writer.writerow(["Events", "", *row.values()])
    for row in data["shots"]:
        writer.writerow(["Shots", "", *row.values()])
    return output.getvalue().encode("utf-8")


def render_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    for row in data["overview"].items():
        overview.append(list(row))
    for title, rows in (("Players", data["players"]), ("Events", data["events"]), ("Shots", data["shots"])):
        sheet = workbook.create_sheet(title)
        if rows:
            sheet.append(list(rows[0].keys()))
            for row in rows:
                sheet.append(list(row.values()))
        else:
            sheet.append(["No data"])
    tactical = workbook.create_sheet("Tactical")
    tactical.append(["Team", "Formation", "Width", "Depth", "Compactness"])
    for row in data["tactical"]:
        tactical.append(list(row.values()))
    passing = workbook.create_sheet("Passing Network")
    passing.append(["Team", "Source", "Target", "Passes"])
    for row in data["passing"]:
        passing.append(list(row.values()))
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def render_pdf(data: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    stream = io.BytesIO()
    document = SimpleDocTemplate(stream, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph("Football Match Report", styles["Title"]), Spacer(1, 8)]
    for title, rows in (("Overview", [[key, value] for key, value in data["overview"].items()]), ("Players", [list(row.values()) for row in data["players"]]), ("Events", [list(row.values()) for row in data["events"]]), ("Shots", [list(row.values()) for row in data["shots"]]), ("Tactical", [list(row.values()) for row in data["tactical"]]), ("Passing Network", [list(row.values()) for row in data["passing"]])):
        story.extend([Paragraph(title, styles["Heading2"])])
        table_rows = rows or [["No data"]]
        table = Table(table_rows, repeatRows=1 if rows else 0)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9ead3")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
        ]))
        story.extend([table, Spacer(1, 8)])
    document.build(story)
    return stream.getvalue()